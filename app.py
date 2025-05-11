
from flask import Flask, render_template, request
from openpyxl import load_workbook
import os
import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

app = Flask(__name__)

TEMPLATE_PATH = "Datos_Completados.xlsx"
MODIFIED_PATH = "Archivo_Modificado.xlsx"

# Cargar credenciales desde variable de entorno
if os.getenv("GOOGLE_CREDS_JSON"):
    with open("client_secrets.json", "w") as f:
        f.write(os.getenv("GOOGLE_CREDS_JSON"))

CREDENTIALS_FILE = "client_secrets.json"
DRIVE_FOLDER_ID = os.getenv("DRIVE_FOLDER_ID")

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    dato1 = request.form["dato1"]
    dato2 = request.form["dato2"]
    dato3 = request.form["dato3"]

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws["B3"] = float(dato1)
    ws["B4"] = float(dato2)
    ws["B5"] = float(dato3)
    wb.save(MODIFIED_PATH)

    upload_to_drive(MODIFIED_PATH)
    return "Archivo subido exitosamente a Google Drive."

def upload_to_drive(file_path):
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_FILE,
        scopes=["https://www.googleapis.com/auth/drive.file"],
    )
    service = build("drive", "v3", credentials=creds)
    file_metadata = {
        "name": f"Excel_Modificado_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "parents": [DRIVE_FOLDER_ID]
    }
    media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service.files().create(body=file_metadata, media_body=media, fields="id").execute()

if __name__ == "__main__":
    app.run(debug=True)
